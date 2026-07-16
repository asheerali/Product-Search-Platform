"""
XLSX parser — extracts:
  • each sheet as a list of row dicts (first row = header)
  • embedded images from sheets saved to pics/
"""
import logging
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class ParsedSheet:
    sheet_name: str
    headers: list[str]
    rows: list[dict]       # each row is {header: value}
    image_paths: list[str] = field(default_factory=list)


@dataclass
class XLSXParseResult:
    document_id: str
    filename: str
    sheets: list[ParsedSheet] = field(default_factory=list)

    @property
    def all_rows(self) -> list[dict]:
        rows = []
        for sheet in self.sheets:
            for row in sheet.rows:
                rows.append({"_sheet": sheet.sheet_name, **row})
        return rows

    @property
    def all_image_paths(self) -> list[str]:
        paths = []
        for s in self.sheets:
            paths.extend(s.image_paths)
        return paths


def _clean_value(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def parse_xlsx(file_path: str, document_id: str, pics_base_dir: str) -> XLSXParseResult:
    import openpyxl

    from app.services.ingestion.image_extractor import save_extracted_image

    path = Path(file_path)
    result = XLSXParseResult(document_id=document_id, filename=path.name)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            if not rows_data:
                continue

            # Find the first non-empty row as the header
            header_row = None
            header_idx = 0
            for idx, row in enumerate(rows_data):
                if any(cell is not None for cell in row):
                    header_row = [_clean_value(c) or f"col_{i}" for i, c in enumerate(row)]
                    header_idx = idx
                    break

            if header_row is None:
                continue

            parsed_rows = []
            for row in rows_data[header_idx + 1:]:
                if all(c is None for c in row):
                    continue  # skip blank rows
                row_dict = {header_row[i]: _clean_value(v) for i, v in enumerate(row) if i < len(header_row)}
                if any(v for v in row_dict.values()):
                    parsed_rows.append(row_dict)

            # Extract embedded images
            image_paths = []
            if hasattr(ws, '_images'):
                for img_idx, img in enumerate(ws._images):
                    try:
                        image_data = img._data()
                        source_ref = f"sheet_{sheet_name}_img_{img_idx}"
                        saved_path = save_extracted_image(
                            image_bytes=image_data,
                            ext="png",
                            document_id=document_id,
                            source_ref=source_ref,
                            pics_base_dir=pics_base_dir,
                        )
                        if saved_path:
                            image_paths.append(saved_path)
                    except Exception as e:
                        logger.warning("Could not extract image from sheet %s: %s", sheet_name, e)

            result.sheets.append(
                ParsedSheet(
                    sheet_name=sheet_name,
                    headers=header_row,
                    rows=parsed_rows,
                    image_paths=image_paths,
                )
            )

        wb.close()
    except Exception as e:
        logger.error("XLSX parsing failed for %s: %s", file_path, e)

    return result
